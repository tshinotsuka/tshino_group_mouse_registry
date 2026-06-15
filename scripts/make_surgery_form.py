#!/usr/bin/env python3
"""make_surgery_form.py — IBL Surgery フォーム生成 (registry → docx)

データ源は **ローカル xlsx** でも **Google Sheets 直読み** でも可。

使い方:
  # ローカル xlsx から
  py make_surgery_form.py registry.xlsx pr-ts301-01

  # Google Sheets から直読み（サービスアカウント認証）
  py make_surgery_form.py "<SheetID か 共有URL>" pr-ts301-01 --creds key.json

  # テンプレ / 出力先を指定
  py make_surgery_form.py registry.xlsx pr-ts301-01 \
      --template templates/SurgeryInformation_template.docx --out out.docx

依存: docxtpl(必須), openpyxl(xlsx時), gspread(Google Sheets時)
  py -m pip install docxtpl openpyxl gspread

充填欄: mouse_id / sex / age_weeks(手術時) / surgery_date / anesthesia /
        surgery_type / site1(region,AP,ML,DV)
手書きのまま残す欄: 体重・時刻・チェックリスト・鎮痛・R/L丸囲み・site2座標
"""
import argparse, os, sys, datetime

TABLES = ("procedures", "animals", "colony")


def _s(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    return str(v).strip()


def load_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in TABLES:
        if name not in wb.sheetnames:
            out[name] = []
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            out[name] = []
            continue
        hdr = [h.strip() if isinstance(h, str) else h for h in rows[0]]
        recs = []
        for r in rows[1:]:
            recs.append({hdr[i]: (r[i] if i < len(r) else None)
                         for i in range(len(hdr)) if hdr[i]})
        out[name] = recs
    return out


def load_gsheet(sheet, creds):
    import gspread
    gc = gspread.service_account(filename=creds)
    sh = gc.open_by_url(sheet) if str(sheet).startswith("http") else gc.open_by_key(sheet)
    out = {}
    for name in TABLES:
        try:
            out[name] = sh.worksheet(name).get_all_records()
        except Exception:
            out[name] = []
    return out


def index_by(recs, key):
    d = {}
    for r in recs:
        k = _s(r.get(key))
        if k:
            d.setdefault(k, r)
    return d


def to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def fmt_date(v):
    d = to_date(v)
    return d.strftime("%Y/%m/%d") if d else _s(v)


def build_context(tables, pid, age_from="auto"):
    procs = index_by(tables["procedures"], "procedure_id")
    if pid not in procs:
        sys.exit(f"[error] procedure_id '{pid}' が procedures に見つかりません")
    p = procs[pid]
    mouse = _s(p.get("mouse_id"))
    a = index_by(tables["animals"], "mouse_id").get(mouse, {})
    bid = _s(a.get("birth_id"))
    c = index_by(tables["colony"], "birth_id").get(bid, {}) if bid else {}

    # 手術時週齢: colony.dob → animals.dob_intake の順に dob を採用し procedure_date から算出
    age = ""
    if age_from != "none":
        pdate = to_date(p.get("procedure_date"))
        if age_from == "colony":
            dob = to_date(c.get("dob"))
        elif age_from == "animals":
            dob = to_date(a.get("dob_intake"))
        else:
            dob = to_date(c.get("dob")) or to_date(a.get("dob_intake"))
        if pdate and dob:
            age = round((pdate - dob).days / 7, 1)
    if age == "":
        age = _s(a.get("age_weeks"))  # フォールバック(at-sacの可能性あり)

    g = lambda d, k: _s(d.get(k))

    def _join(drug, route):
        a = str(g(p, drug)) if g(p, drug) != "" else ""
        b = str(g(p, route)) if g(p, route) != "" else ""
        return f"{a}, {b}" if (a and b) else (a or b)

    ctx = dict(
        mouse_id=mouse,
        sex=g(a, "sex"),
        age_weeks=age,
        surgery_date=fmt_date(p.get("procedure_date")),
        anesthesia=_join("anesthesia", "anesthesia_route"),
        analgesia=_join("analgesia", "analgesia_route"),
        surgery_type=g(p, "procedure_type"),
        site1_region=g(p, "site1_region"),
        site1_ap=g(p, "site1_ap"),
        site1_ml=g(p, "site1_ml"),
        site1_dv=g(p, "site1_dv"),
    )
    return {k: ("" if v == "" else str(v)) for k, v in ctx.items()}, p


def main():
    ap = argparse.ArgumentParser(description="IBL Surgery フォーム生成 (xlsx / Google Sheets)")
    ap.add_argument("source", help="registry.xlsx か Google Sheet ID/共有URL")
    ap.add_argument("procedure_id")
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--template",
                    default=os.path.join(here, "..", "templates", "SurgeryInformation_template.docx"))
    ap.add_argument("--out", default=None)
    ap.add_argument("--creds", default=None, help="gspread サービスアカウント json (Google Sheets時必須)")
    ap.add_argument("--age-from", choices=["auto", "colony", "animals", "none"], default="auto")
    a = ap.parse_args()

    if a.source.lower().endswith((".xlsx", ".xlsm")):
        if not os.path.exists(a.source):
            sys.exit(f"[error] ファイルが見つかりません: {a.source}\n"
                     f"        ダウンロードした xlsx の実際の場所をパスで指定してください。\n"
                     f"        例: py make_surgery_form.py \"%USERPROFILE%\\Downloads\\registry.xlsx\" {a.procedure_id}")
        tables = load_xlsx(a.source)
    elif a.creds:
        tables = load_gsheet(a.source, a.creds)
    else:
        sys.exit("[error] xlsx のパスを渡すか、Google Sheets の場合は --creds <service_account.json> を指定してください")

    ctx, _ = build_context(tables, a.procedure_id, a.age_from)
    if not os.path.exists(a.template):
        sys.exit(f"[error] テンプレートが見つかりません: {a.template}")

    from docxtpl import DocxTemplate
    tpl = DocxTemplate(a.template)
    tpl.render(ctx)
    out = a.out or os.path.join(here, "..", "output", f"surgery_{a.procedure_id}.docx")
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    tpl.save(out)

    print("生成:", out)
    print("  充填:", {k: v for k, v in ctx.items() if v != ""})
    miss = [k for k in ("sex", "age_weeks", "anesthesia", "surgery_type") if ctx[k] == ""]
    if miss:
        print("  空欄(手書き):", ", ".join(miss))


if __name__ == "__main__":
    main()
