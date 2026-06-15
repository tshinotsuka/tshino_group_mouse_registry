#!/usr/bin/env python3
"""snapshot.py — Google Sheets の xlsx 書き出しを CSV スナップショットに固定し、
整合チェックを通してから git commit するためのツール。

パイプライン:
  ① xlsx を読む  ② snapshots/<table>.csv に書き出す
  ③ check_registry.py を CSV に対して実行（= ゲート。PASS しなければ commit しない）
  ④ --commit 指定かつ PASS のとき git add snapshots/ && git commit

使い方:
  # 書き出し + チェックのみ（commit しない）
  python snapshot.py "C:\\Users\\takan\\Downloads\\tshino_group_mouse_registry.xlsx"

  # チェックを通したら commit まで
  python snapshot.py "...xlsx" --commit -m "snapshot before analysis run"

  # warning も失格扱いにする
  python snapshot.py "...xlsx" --strict

仕様:
  - 出力タブ: lines, colony, matings, litters, animals, enrollments, projects, procedures, vocab
    （dashboard は派生のため出力しない）
  - TODAY() を含む揮発フォーミュラ列（例: colony.age_weeks）は自動検出して CSV から除外
    （毎回値が変わり diff が汚れるため。値は dob から再計算できる）
  - 行順は Sheet のまま（挿入順）/ 日付は YYYY-MM-DD / 整数は .0 を付けない
  - 依存: openpyxl（pandas は check_registry 側）
"""
import argparse, csv, datetime, os, subprocess, sys

KNOWN = ["lines", "colony", "matings", "litters", "animals",
         "enrollments", "projects", "procedures", "vocab"]


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v)


def volatile_cols(ws_f, headers):
    """ヘッダごとに、列内のどこかが TODAY() を含むフォーミュラなら揮発列とみなす。"""
    vol = set()
    rows = list(ws_f.iter_rows(min_row=2, values_only=True))
    for ci, h in enumerate(headers):
        if not h:
            continue
        for r in rows:
            if ci < len(r):
                cell = r[ci]
                if isinstance(cell, str) and "TODAY(" in cell.upper():
                    vol.add(h)
                    break
    return vol


def write_snapshots(xlsx, out_dir):
    import openpyxl
    wb_v = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)   # 値
    wb_f = openpyxl.load_workbook(xlsx, read_only=True, data_only=False)  # 数式
    os.makedirs(out_dir, exist_ok=True)
    name_map = {s.lower(): s for s in wb_v.sheetnames}
    written, skipped = [], []
    for tab in KNOWN:
        if tab not in name_map:
            continue
        ws = wb_v[name_map[tab]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [h.strip() if isinstance(h, str) else h for h in rows[0]]
        vol = volatile_cols(wb_f[name_map[tab]], headers)
        keep = [i for i, h in enumerate(headers) if h and h not in vol]
        out_headers = [headers[i] for i in keep]
        fp = os.path.join(out_dir, f"{tab}.csv")
        with open(fp, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f, lineterminator="\n")
            w.writerow(out_headers)
            for r in rows[1:]:
                cells = [fmt(r[i]) if i < len(r) else "" for i in keep]
                if all(c == "" for c in cells):
                    continue  # 完全空行はスキップ
                w.writerow(cells)
        written.append(tab)
        if vol:
            skipped.append(f"{tab}:{','.join(sorted(vol))}")
    return written, skipped


def run_check(check_py, out_dir, strict):
    cmd = [sys.executable, check_py, out_dir] + (["--strict"] if strict else [])
    print("--- check_registry ---")
    p = subprocess.run(cmd)
    return p.returncode == 0


def git_commit(repo_root, out_dir, message):
    rel = os.path.relpath(out_dir, repo_root)
    print("--- git commit ---")
    add = subprocess.run(["git", "-C", repo_root, "add", rel])
    if add.returncode != 0:
        print("git add に失敗（このフォルダは git リポジトリ？）")
        return False
    # 変更が無ければ commit しない
    diff = subprocess.run(["git", "-C", repo_root, "diff", "--cached", "--quiet"])
    if diff.returncode == 0:
        print("スナップショットに変更なし（commit 不要）")
        return True
    c = subprocess.run(["git", "-C", repo_root, "commit", "-m", message])
    return c.returncode == 0


def main():
    ap = argparse.ArgumentParser(description="registry snapshot (xlsx -> CSV -> check -> commit)")
    ap.add_argument("xlsx", help="Google Sheets を書き出した .xlsx")
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--snapshots-dir", default=os.path.join(here, "..", "snapshots"))
    ap.add_argument("--strict", action="store_true", help="warning も失格扱い")
    ap.add_argument("--no-check", action="store_true", help="check をスキップ（非推奨）")
    ap.add_argument("--commit", action="store_true", help="PASS なら git commit まで実行")
    ap.add_argument("-m", "--message", default=None)
    a = ap.parse_args()

    if not os.path.exists(a.xlsx):
        sys.exit(f"[error] ファイルが見つかりません: {a.xlsx}")
    out_dir = os.path.abspath(a.snapshots_dir)

    written, skipped = write_snapshots(a.xlsx, out_dir)
    print(f"書き出し: {len(written)} tables -> {out_dir}")
    print("  " + ", ".join(written))
    if skipped:
        print("  揮発列を除外:", "; ".join(skipped))

    ok = True
    if not a.no_check:
        check_py = os.path.join(here, "check_registry.py")
        ok = run_check(check_py, out_dir, a.strict)
        if not ok:
            print("\n[gate] check が通りませんでした → commit しません。")
            print("  ・上に 'RESULT: FAIL' が出ている場合: データ不整合。Sheet を直して再エクスポート。")
            print("  ・Traceback / ModuleNotFoundError の場合: check の実行エラー。")
            print("    例) pandas 未導入 → `pip install pandas` を実行してください。")
            sys.exit(1)

    if a.commit:
        repo_root = os.path.dirname(out_dir)
        msg = a.message or f"snapshot {datetime.datetime.now():%Y-%m-%d %H:%M}"
        if git_commit(repo_root, out_dir, msg):
            print("commit 完了:", msg)
        else:
            sys.exit("[error] git commit に失敗")
    else:
        print("\n(--commit 未指定: CSV 更新のみ。確認後 `git add snapshots && git commit` か、--commit で再実行)")


if __name__ == "__main__":
    main()
