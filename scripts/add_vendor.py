#!/usr/bin/env python3
"""add_vendor.py — register vendor-purchased mice into the colony (v1.3).

For animals that come in from a vendor (not bred in-house). Unlike add_litter.py
there is NO litter: no sire/dam, no litter_id, no matings row. Rows are appended
at the bottom with source=vendor and auto-assigned consecutive birth_ids
(next = max existing number in that line + 1). Genotype is usually known on arrival,
so pass --gt-* and --genotype-status confirmed (WT lines have no assays -> use na).

Examples:
  # 5 WT females from CLEA
  python add_vendor.py registry.xlsx --line wt --females 5 --dob 2026-06-01 \
      --vendor CLEA --genotype-status na
  # 3 hemizygous 5xFAD males from Jackson (genotype confirmed by vendor)
  python add_vendor.py registry.xlsx --line 5xfad --males 3 --dob 2026-05-20 \
      --vendor Jackson --gt-app=+/- --gt-ps1=+/- --genotype-status confirmed
  python add_vendor.py registry.xlsx --line wt --females 5 --dob 2026-06-01 --vendor CLEA --dry-run
"""
import argparse, datetime as dt, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE=Font(name="Arial"); VEND=PatternFill("solid",fgColor="FDE9D9"); GT=PatternFill("solid",fgColor="EDE7F6")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
COLS=["birth_id","line_id","number","sex","dob","age_weeks","source","litter_id","sire","dam",
      "generation","gt_app","gt_ps1","gt_aqp4","gt_cagegfp","genotype","genotype_status","status","status_date","notes"]

def col_index(ws,name):
    for c in range(1,ws.max_column+1):
        if str(ws.cell(1,c).value).strip()==name: return c
    raise SystemExit(f"column '{name}' not found")
def last_row(ws,kc=1):
    r=ws.max_row
    while r>=2 and (ws.cell(r,kc).value in (None,"")): r-=1
    return r
def max_number(ws,line,nc,lc):
    mx=0
    for r in range(2,ws.max_row+1):
        if str(ws.cell(r,lc).value).strip()==line:
            v=ws.cell(r,nc).value
            if isinstance(v,(int,float)): mx=max(mx,int(v))
    return mx

def main(argv):
    ap=argparse.ArgumentParser()
    ap.add_argument("xlsx"); ap.add_argument("--line",required=True)
    ap.add_argument("--males",type=int,default=0); ap.add_argument("--females",type=int,default=0); ap.add_argument("--unknown",type=int,default=0)
    ap.add_argument("--dob",required=True,help="DOB (or estimated DOB from vendor age)")
    ap.add_argument("--vendor",default="",help="vendor name (CLEA/Jackson/RIKEN...) -> notes")
    ap.add_argument("--gt-app",default=""); ap.add_argument("--gt-ps1",default=""); ap.add_argument("--gt-aqp4",default=""); ap.add_argument("--gt-cagegfp",default="")
    ap.add_argument("--genotype-status",default="confirmed"); ap.add_argument("--status",default="alive")
    ap.add_argument("--generation",default=""); ap.add_argument("--notes",default=""); ap.add_argument("--dry-run",action="store_true")
    a=ap.parse_args(argv)
    dob=dt.datetime.strptime(a.dob,"%Y-%m-%d").date()
    sexes=["M"]*a.males+["F"]*a.females+["U"]*a.unknown; n=len(sexes)
    if n==0: raise SystemExit("set --males/--females/--unknown")
    wb=load_workbook(a.xlsx); C=wb["colony"]; ci={h:col_index(C,h) for h in COLS}
    GL={m:get_column_letter(ci[m]) for m in ("gt_app","gt_ps1","gt_aqp4","gt_cagegfp")}
    def geno(r):
        a1,b1,c1,d1=GL["gt_app"],GL["gt_ps1"],GL["gt_aqp4"],GL["gt_cagegfp"]
        return (f'=MID(IF({a1}{r}="","","; APP:"&{a1}{r})&IF({b1}{r}="","","; PS1:"&{b1}{r})'
                f'&IF({c1}{r}="","","; AQP4:"&{c1}{r})&IF({d1}{r}="","","; CAG-EGFP:"&{d1}{r}),3,200)')
    start=max_number(C,a.line,ci["number"],ci["line_id"])+1; base=last_row(C,ci["birth_id"])+1
    note=" ".join(x for x in [(f"vendor:{a.vendor}" if a.vendor else "vendor"),f"recv:{a.dob}",a.notes] if x)
    assigned=[]
    for i,sx in enumerate(sexes):
        num=start+i; bid=f"{a.line}-{num}"; assigned.append((bid,sx))
        if a.dry_run: continue
        r=base+i
        vals={"birth_id":bid,"line_id":a.line,"number":num,"sex":sx,"dob":dob,"source":"vendor",
              "litter_id":"","sire":"","dam":"","generation":(int(a.generation) if str(a.generation).strip() else ""),
              "gt_app":a.gt_app,"gt_ps1":a.gt_ps1,"gt_aqp4":a.gt_aqp4,"gt_cagegfp":a.gt_cagegfp,
              "genotype_status":a.genotype_status,"status":a.status,"status_date":dt.date.today(),"notes":note[:80]}
        for h,col in ci.items():
            if h=="age_weeks": cell=C.cell(r,col,value=f'=IF(E{r}="","",ROUND((TODAY()-E{r})/7,1))')
            elif h=="genotype": cell=C.cell(r,col,value=geno(r))
            else: cell=C.cell(r,col,value=vals.get(h,""))
            cell.font=BASE; cell.border=BORD
        C.cell(r,ci["birth_id"]).fill=VEND
        for m in ("gt_app","gt_ps1","gt_aqp4","gt_cagegfp"): C.cell(r,ci[m]).fill=GT
        C.cell(r,ci["dob"]).number_format="yyyy-mm-dd"; C.cell(r,ci["status_date"]).number_format="yyyy-mm-dd"
    print(f"vendor intake: line={a.line} n={n} ({a.males}M/{a.females}F"+(f"/{a.unknown}U" if a.unknown else "")+f") source=vendor {a.vendor}")
    print(f"  birth_id: {assigned[0][0]} .. {assigned[-1][0]}")
    for bid,sx in assigned: print(f"    {bid}\t{sx}")
    if a.dry_run: print("  [dry-run] nothing written. (no matings row created for vendor intake)")
    else:
        wb.save(a.xlsx); print(f"  next {a.line} number now {start+n}. No matings row (vendor). Saved.")
        print("  -> run check_registry.py before analysis.")
    return 0

if __name__=="__main__": sys.exit(main(sys.argv[1:]))
