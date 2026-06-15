#!/usr/bin/env python3
"""add_litter.py — register a newly born/delivered litter into the colony (v1.2).

Auto-assigns consecutive birth_ids (next = max existing number + 1), appends one
colony row per pup (males, then females, then unknown), and (for bred litters) a
matings litter row. Per-assay genotype columns (gt_app/gt_ps1/gt_aqp4/gt_cagegfp) are
left blank by default (pups = genotype_status pending); set them later in the Sheet,
or pass --gt-* if known at intake. The combined `genotype` column is a formula.

Examples:
  python add_litter.py registry.xlsx --line aqp4ko --males 6 --females 4 \
      --dob 2026-06-10 --sire aqp4ko-122 --dam aqp4ko-123 --generation 5
  python add_litter.py registry.xlsx --line wt --females 3 --dob 2026-06-01 \
      --source vendor --status alive --genotype-status na --no-mating
  python add_litter.py registry.xlsx --line aqp4ko --males 6 --females 4 --dob 2026-06-10 --dry-run
"""
import argparse, datetime as dt, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE=Font(name="Arial"); MINE=PatternFill("solid",fgColor="E2EFDA"); GT=PatternFill("solid",fgColor="EDE7F6")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
COLS=["birth_id","line_id","number","sex","dob","age_weeks","source","litter_id","sire","dam",
      "generation","gt_app","gt_ps1","gt_aqp4","gt_cagegfp","genotype","genotype_status","status","status_date","notes"]

def col_index(ws,name):
    for c in range(1,ws.max_column+1):
        if str(ws.cell(1,c).value).strip()==name: return c
    raise SystemExit(f"column '{name}' not found in {ws.title}")
def last_row(ws,kc=1):
    r=ws.max_row
    while r>=2 and (ws.cell(r,kc).value in (None,"")): r-=1
    return r
def max_number(ws,line,nc,lc):
    mx=0
    for r in range(2,ws.max_row+1):
        if str(ws.cell(r,lc).value).strip()==line:
            v=ws.cell(r,nc).value
            if isinstance(v,(int,float)): mx=max(mx,int(v))
    return mx

def main(argv):
    ap=argparse.ArgumentParser()
    ap.add_argument("xlsx"); ap.add_argument("--line",required=True)
    ap.add_argument("--males",type=int,default=0); ap.add_argument("--females",type=int,default=0)
    ap.add_argument("--unknown",type=int,default=0)
    ap.add_argument("--dob",required=True); ap.add_argument("--sire",default=""); ap.add_argument("--dam",default="")
    ap.add_argument("--generation",default="")
    ap.add_argument("--gt-app",default=""); ap.add_argument("--gt-ps1",default=""); ap.add_argument("--gt-aqp4",default=""); ap.add_argument("--gt-cagegfp",default="")
    ap.add_argument("--genotype-status",default="pending"); ap.add_argument("--status",default="alive")
    ap.add_argument("--source",default="in_house"); ap.add_argument("--litter",default=""); ap.add_argument("--mating",default="")
    ap.add_argument("--no-mating",action="store_true"); ap.add_argument("--notes",default=""); ap.add_argument("--dry-run",action="store_true")
    a=ap.parse_args(argv)
    dob=dt.datetime.strptime(a.dob,"%Y-%m-%d").date(); dc=a.dob.replace("-","")
    sexes=["M"]*a.males+["F"]*a.females+["U"]*a.unknown; n=len(sexes)
    if n==0: raise SystemExit("set --males/--females/--unknown")
    wb=load_workbook(a.xlsx); C=wb["colony"]; ci={h:col_index(C,h) for h in COLS}
    GTL={m:get_column_letter(ci[m]) for m in ("gt_app","gt_ps1","gt_aqp4","gt_cagegfp")}
    start=max_number(C,a.line,ci["number"],ci["line_id"])+1
    litter=a.litter or f"{a.line}-l{dc}"
    def geno(r):
        a,b,c,d=GTL["gt_app"],GTL["gt_ps1"],GTL["gt_aqp4"],GTL["gt_cagegfp"]
        return (f'=MID(IF({a}{r}="","","; APP:"&{a}{r})&IF({b}{r}="","","; PS1:"&{b}{r})'
                f'&IF({c}{r}="","","; AQP4:"&{c}{r})&IF({d}{r}="","","; CAG-EGFP:"&{d}{r}),3,200)')
    assigned=[]; base=last_row(C,ci["birth_id"])+1
    for i,sx in enumerate(sexes):
        num=start+i; bid=f"{a.line}-{num}"; assigned.append((bid,sx))
        if a.dry_run: continue
        r=base+i
        vals={"birth_id":bid,"line_id":a.line,"number":num,"sex":sx,"dob":dob,"source":a.source,
              "litter_id":litter,"sire":a.sire,"dam":a.dam,
              "generation":(int(a.generation) if str(a.generation).strip() else ""),
              "gt_app":a.gt_app,"gt_ps1":a.gt_ps1,"gt_aqp4":a.gt_aqp4,"gt_cagegfp":a.gt_cagegfp,
              "genotype_status":a.genotype_status,"status":a.status,"status_date":dt.date.today(),"notes":a.notes}
        for h,col in ci.items():
            if h=="age_weeks": cell=C.cell(r,col,value=f'=IF(E{r}="","",ROUND((TODAY()-E{r})/7,1))')
            elif h=="genotype": cell=C.cell(r,col,value=geno(r))
            else: cell=C.cell(r,col,value=vals.get(h,""))
            cell.font=BASE; cell.border=BORD
        C.cell(r,ci["birth_id"]).fill=MINE
        for m in ("gt_app","gt_ps1","gt_aqp4","gt_cagegfp"): C.cell(r,ci[m]).fill=GT
        C.cell(r,ci["dob"]).number_format="yyyy-mm-dd"; C.cell(r,ci["status_date"]).number_format="yyyy-mm-dd"
    made=False
    if not a.no_mating and a.sire and a.dam and "matings" in wb.sheetnames and not a.dry_run:
        M=wb["matings"]; mi={h:col_index(M,h) for h in ["mating_id","line_id","sire_id","dam_id","pair_start_date","litter_dob","litter_id","n_pups","n_male","n_female","status","notes"]}
        mr=last_row(M,mi["mating_id"])+1; mating=a.mating or f"{a.line}-m{dc}"
        mv={"mating_id":mating,"line_id":a.line,"sire_id":a.sire,"dam_id":a.dam,"litter_dob":dob,"litter_id":litter,
            "n_pups":n,"n_male":a.males,"n_female":a.females+a.unknown,"status":"littered","notes":a.notes}
        for h,col in mi.items(): c=M.cell(mr,col,value=mv.get(h,"")); c.font=BASE; c.border=BORD
        M.cell(mr,mi["litter_dob"]).number_format="yyyy-mm-dd"; M.cell(mr,mi["mating_id"]).fill=MINE; made=True
    print(f"line={a.line} pups={n} ({a.males}M/{a.females}F"+(f"/{a.unknown}U" if a.unknown else "")+")")
    print(f"  birth_id: {assigned[0][0]} .. {assigned[-1][0]}")
    for bid,sx in assigned: print(f"    {bid}\t{sx}")
    print(f"  litter_id: {litter}"+("  (+matings row)" if made else ""))
    if a.dry_run: print("  [dry-run] nothing written.")
    else:
        wb.save(a.xlsx); print(f"  next {a.line} number now {start+n}. genotyping pending (set gt_* later). Saved.")
        print("  -> run check_registry.py before analysis.")
    return 0

if __name__=="__main__": sys.exit(main(sys.argv[1:]))
